# from django.shortcuts import render, redirect, get_object_or_404
# from .models import Doctor, Patient, Appointment
# from django.contrib.auth.decorators import login_required
# from .forms import AppointmentForm
#
# def home(request):
#     return render(request, 'appointments/index.html')
#
# def doctor_list(request):
#     doctors = Doctor.objects.all()
#     return render(request, 'appointments/doctors.html', {'doctors': doctors})
#
# @login_required
# def make_appointment(request, doctor_id):
#     doctor = get_object_or_404(Doctor, id=doctor_id)
#     patient, created = Patient.objects.get_or_create(user=request.user)
#
#     if request.method == 'POST':
#         form = AppointmentForm(request.POST)
#         if form.is_valid():
#             appointment = form.save(commit=False)
#             appointment.doctor = doctor
#             appointment.patient = patient
#             appointment.save()
#             return redirect('doctor_list')
#     else:
#         form = AppointmentForm()
#
#     return render(request, 'appointments/make_appointment.html', {'form': form, 'doctor': doctor})
from django.http import FileResponse, Http404
from django.shortcuts import render, get_object_or_404
from django.conf import settings
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook
import os
from .models import Patient, Appointment
from .forms import AppointmentForm
from django.contrib import messages
from .models import Doctor, Specialization
from django.contrib.auth.decorators import login_required
from .models import ContactRequest

def home(request):
    return render(request, 'appointments/index.html')
def doctors_list(request):
    specialization=request.GET.get('specialization')
    if specialization:
        doctors=Doctor.objects.filter(specialization__iexact=specialization)
    else:
        doctors=Doctor.objects.all()
    return render(request, 'doctors_list.html', {'doctors': doctors})


def doctor_detail(request, doctor_id):
    doctor=get_object_or_404(Doctor, id=doctor_id)
    if request.method == 'POST':
        form=AppointmentForm(request.POST)
        if form.is_valid():
            appointment_time=form.cleaned_data['appointment_time']
            # Проверяем, свободно ли время
            existing_appointment=Appointment.objects.filter(doctor=doctor, appointment_time=appointment_time).exists()
            if existing_appointment:
                messages.error(request, 'Таңдалған уақыт бос емес. Басқа уақытқа қойыңыз.')
            else:
                # Создаем или находим пациента
                patient, created=Patient.objects.get_or_create(
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    phone_number=form.cleaned_data['phone_number'],
                )
                # Создаем запись
                Appointment.objects.create(
                    doctor=doctor,
                    patient=patient,
                    appointment_time=appointment_time,
                    description=form.cleaned_data['description'],
                )
                messages.success(request, 'Жазылым сәтті аяқталды!')
                return redirect('appointment_confirmation')
    else:
        form=AppointmentForm()

    # Список уже занятых времен для отображения
    booked_times=Appointment.objects.filter(doctor=doctor).values_list('appointment_time', flat=True)

    return render(request, 'doctor_detail.html', {
        'doctor': doctor,
        'form': form,
        'booked_times': booked_times,
    })


@login_required
def make_appointment(request, doctor_id):
    doctor = get_object_or_404(Doctor, id=doctor_id)

    if request.method == 'POST':
        appointment_time = request.POST.get('appointment_time')
        description = request.POST.get('description')

        try:
            patient = request.user.patient
        except Patient.DoesNotExist:
            messages.error(request, 'Қате: сіздің аккаунтыңыз пациент ретінде тіркелмеген.')
            return redirect('dashboard')

        Appointment.objects.create(
            doctor=doctor,
            patient=patient,
            appointment_time=appointment_time,
            description=description
        )

        messages.success(request, 'Сіз сәтті түрде жазылдыңыз!')
        return redirect('dashboard')

    return render(request, 'appointments/make_appointment.html', {'doctor': doctor})
def doctors_by_specialization(request, spec_id):
    specialization = Specialization.objects.get(id=spec_id)
    doctors = Doctor.objects.filter(specialization=specialization)
    return render(request, 'appointments/doctors_by_specialization.html', {
        'specialization': specialization,
        'doctors': doctors
    })





import os
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import ContactRequest

def contact(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")

        if name and phone:
            ContactRequest.objects.create(name=name, phone=phone)
            messages.success(request, "Сәтті жіберілді.")
        else:
            messages.error(request, "Барлық өрістерді толтырыңыз.")

        return redirect('contact')  # URL name

    return render(request, 'appointments/contact_form.html')

# def contact(request):
#     if request.method == "POST":
#         name = request.POST.get("name")
#         phone = request.POST.get("phone")
#
#         if not name or not phone:
#             messages.error(request, "Барлық өрістерді толтырыңыз.")
#             return redirect('/')
#         os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
#
#         try:
#             if not os.path.exists(EXCEL_FILE):
#                 wb = Workbook()
#                 ws = wb.active
#                 ws.append(["Аты-жөні", "Телефон нөмірі"])
#             else:
#                 wb = load_workbook(EXCEL_FILE)
#                 ws = wb.active
#
#             ws.append([name, phone])
#             wb.save(EXCEL_FILE)
#
#             messages.success(request, "Сәтті жіберілді. Деректер Excel-ге сақталды.")
#         except Exception as e:
#             messages.error(request, f"Қате: {e}")
#
#         return redirect('/')
#
#     return render(request, 'appointments/index.html')

def download_exel(request):
    file_path = os.path.join(settings.MEDIA_ROOT, 'contacts', 'contact_data.xlsx')
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='contact_data.xlsx')
    else:
        raise Http404("Файл табылмады ")