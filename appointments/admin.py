
from .models import Doctor, Patient, Appointment, Specialization
from django.contrib import admin
from .models import ContactRequest
from openpyxl import Workbook, load_workbook
import os
from django.conf import settings

admin.site.register(Doctor)
admin.site.register(Patient)
admin.site.register(Appointment)
admin.site.register(Specialization)





EXCEL_FILE = os.path.join(settings.MEDIA_ROOT, 'contacts', 'contact_data.xlsx')


@admin.register(ContactRequest)
class ContactRequestAdmin(admin.ModelAdmin):
    list_display = ("name", "phone", "created_at")
    actions = ["export_to_excel"]

    @admin.action(description="Exelге экспорттау")
    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.append(["Аты-жөні", "Телефон нөмірі", "Жіберілген уақыты"])
        for contact in queryset:
            ws.append([contact.name, contact.phone, contact.created_at.strftime('%d.%m.%Y %H:%M')])
        from django.http import HttpResponse
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=contacts.xlsx'
        wb.save(response)
        return response

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)

        os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(["Аты-жөні", "Телефон нөмірі", "Жіберілген уақыты"])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        ws.append([obj.name, obj.phone, obj.created_at.strftime('%d.%m.%Y %H:%M')])
        wb.save(EXCEL_FILE)

